import os, sys
import requests
import re
from openpyxl import load_workbook, Workbook
import ConfigParser
from time import sleep
import wx
import threading

class myThread(threading.Thread): #separate worker thread for unblocked ui
    
    def __init__(self, appObj): #appObj contains dialog ui elements
        threading.Thread.__init__(self)
        self.appObj = appObj
        self.Terminate = False
        
    def quit(self):
        if self.isAlive():
            self.Terminate = True
            
    def log(self, message):
        self.appObj.tc4.AppendText(message)

    def run(self):
        if self.Terminate: return
        try:
            cookies = dict()
            headers = dict()

            self.log('\nLoading Configuration File...')
            
            config = ConfigParser.ConfigParser()
            config.read(self.appObj.tc1.Value) #get configuration
            
            main = dict(config.items('Main')) #main config
            get = dict(config.items('Get')) #first stage: scrap data and store it in variables
            prefvars = dict(config.items('Generate')) #second stage: process variables stored
            put = dict(config.items('Put')) #third stage: store processed variables in excel sheet
            
            if main['require_login'] == 'yes':
                self.log('\nAttempting to Login...')
                
                login = dict(config.items('Login'))
                exec "import " + login['module'] #import login module
                exec "cookies = " + login['function'] #function returns auth cookies in login module
                
                if not cookies:
                    self.log("\nERROR: Login failed.. Exiting..")
                    return
                else:
                    self.log('\nSuccessfully logged in.. \nCookie:' + str(cookies))
                                
            headers["User-Agent"] = "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.17 (KHTML, like Gecko) Chrome/24.0.1312.60 Safari/537.17"
            
            self.log("\nReading source file for keys..")
            src_wb = load_workbook(filename=self.appObj.tc2.Value, read_only=True) #load excel workbook
            src_ws = src_wb.worksheets[0] #load 1st worksheet
            
            dest_wb = Workbook() #create new workbook
            dest_ws = dest_wb.active #select 1st worksheet
            
            num_rows = src_ws.max_row #get number of rows
            curr_row = int(main['start_row']) - 1 #current row hack; will be incremented in while loop
            
            while curr_row < num_rows:
                if self.Terminate: return #check abort button has been clicked
                curr_row += 1 #current row hack: actual row :-)

                exec 'keys = dict({0})'.format(main['keys']) #fields to be submitted
                keys_with_value = {}
                for key, value in keys.iteritems():
                    exec 'keys_with_value["{0}"] = src_ws.cell(row=curr_row, column={1}).value'.format(key, str(value)) #for saving keys in excel sheet
                                
                self.log("\nProcessing row: " + str(curr_row))
                
                exec 'additionalvars = dict({0})'.format(main['request_vars']) #additional static vars needs to submitted

                payload = dict(keys_with_value.items() + additionalvars.items()) #complete payload
    
                #print payload
                
                headers["Referer"] = main['referer']
                
                if main['request_type'] == 'post':
                    exec 'req = requests.post("{0}", data=payload, headers=headers, cookies=cookies, verify=False)'.format(main['url']) #post request
                else:
                    exec 'req = requests.get("{0}", params=payload, headers=headers, cookies=cookies, verify=False)'.format(main['url']) #get request
                    
                for k in get:
                    exec '{0}_reg= re.compile({1}, re.M|re.DOTALL)'.format(k, get[k]) #compile regular expression for future use; multi-line and dot matches all named key_reg
                    if main['debug_mode'] == 'true':
                        self.log(req.text) #print complete output if in debug mode
                    exec '{0}_groups = {0}_reg.match(req.text)'.format(k) #grab all groups
                    exec '{0}={0}_groups.group(1) if {0}_groups else ""'.format(k) #grab 1st group and name it as per get config
                    
                for l in prefvars:
                    exec '{0} = {1}'.format(l, prefvars[l]) #process generate step
                
                self.log("\n")
                
                for m in get:
                    exec "self.log({0}\t)".format(m) #print captured variables
                    
                for n in put:
                    if curr_row == 1:
                        dest_ws.cell(row = 1, column = int(n)).value = put[n].upper() #put excel headers
                    act_row = curr_row + 1
                    exec 'dest_ws.cell(row = act_row, column = {0}).value = {1}'.format(n, put[n]) #put excel data; note: differ between exec in this line and for curr_row==1
                    
                dest_wb.save(self.appObj.tc3.Value)
                
                sleep(int(main['sleep'])) #wait for n mseconds

            self.log("\n\nAll right.. Everything went smoothly.. Exiting.. Bye!") #completed
            self.appObj.button1.Enable()
            self.appObj.button2.Enable()
            self.appObj.button3.Enable()
            self.appObj.button4.Enable()
            self.appObj.button5.Disable()
            self.appObj.button6.Enable()
        
        except:
            self.log("\n\n--------------------Error------------------------")
            wx.MessageBox(str(sys.exc_value) + " in Line no: " + str(sys.exc_traceback.tb_lineno), 'Error', wx.OK | wx.ICON_ERROR)
            self.appObj.button1.Enable()
            self.appObj.button2.Enable()
            self.appObj.button3.Enable()
            self.appObj.button4.Enable()
            self.appObj.button5.Disable()
            self.appObj.button6.Enable()

class Hawk(wx.Frame):
    
    def __init__(self, parent, title):
        self.application_path = os.path.dirname(sys.executable)
        super(Hawk, self).__init__(parent, title=title, size=(535, 435), style=wx.SYSTEM_MENU|wx.CAPTION)
        icon1 = wx.Icon(self.application_path + '/Hawk.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(icon1)
        self.InitUI()
        self.Centre()
        self.Show()
        
    def InitUI(self):
        panel = wx.Panel(self)
        sizer = wx.GridBagSizer(5, 5)
        
        text1 = wx.StaticText(panel, label="Python Tool for Websites <saravanakumar.a.o@gmail.com>\n\nCreated by Saravanakumar Arumugam")
        sizer.Add(text1, pos=(0, 0), span=(0, 3), flag=wx.TOP|wx.LEFT|wx.BOTTOM|wx.EXPAND, border=15)
        
        icon = wx.StaticBitmap(panel, label=wx.Bitmap(self.application_path + '/Hawk.png'))
        sizer.Add(icon, pos=(0, 4), flag=wx.TOP|wx.RIGHT|wx.ALIGN_RIGHT, border=5)
        
        line = wx.StaticLine(panel)
        sizer.Add(line, pos=(1, 0), span=(1, 5), 
            flag=wx.EXPAND|wx.BOTTOM, border=10)

        text2 = wx.StaticText(panel, label="Configuration File")
        sizer.Add(text2, pos=(2, 0), flag=wx.LEFT, border=10)

        self.tc1 = wx.TextCtrl(panel, style=wx.TE_READONLY, value="[Select File]")
        sizer.Add(self.tc1, pos=(2, 1), span=(1, 3), flag=wx.TOP|wx.EXPAND)
        
        self.button1 = wx.Button(panel, label="Browse...")
        sizer.Add(self.button1, pos=(2, 4), flag=wx.TOP|wx.RIGHT)
        self.button1.Bind(wx.EVT_BUTTON, self.onConfBrowse)

        text3 = wx.StaticText(panel, label="Source Excel File")
        sizer.Add(text3, pos=(3, 0), flag=wx.LEFT|wx.TOP, border=10)

        self.tc2 = wx.TextCtrl(panel, style=wx.TE_READONLY, value="[Select File]")
        sizer.Add(self.tc2, pos=(3, 1), span=(1, 3), flag=wx.TOP|wx.EXPAND, border=5)

        self.button2 = wx.Button(panel, label="Browse...")
        sizer.Add(self.button2, pos=(3, 4), flag=wx.TOP|wx.RIGHT, border=5)
        self.button2.Bind(wx.EVT_BUTTON, self.onSrcBrowse)

        text4 = wx.StaticText(panel, label="Destination Excel File")
        sizer.Add(text4, pos=(4, 0), flag=wx.TOP|wx.LEFT, border=10)

        self.tc3 = wx.TextCtrl(panel, style=wx.TE_READONLY, value="[Specify File]")
        sizer.Add(self.tc3, pos=(4, 1), span=(1, 3), flag=wx.TOP|wx.EXPAND, border=5)

        self.button3 = wx.Button(panel, label="Browse...")
        sizer.Add(self.button3, pos=(4, 4), flag=wx.TOP|wx.RIGHT, border=5)
        self.button3.Bind(wx.EVT_BUTTON, self.onDestBrowse)

        sb = wx.StaticBox(panel, label="Progress")
        
        boxsizer = wx.StaticBoxSizer(sb, wx.VERTICAL)
        self.tc4 = wx.TextCtrl(panel, value="[Ready]", style=wx.TE_READONLY|wx.TE_MULTILINE, size=(50,100))
        boxsizer.Add(self.tc4, flag=wx.EXPAND|wx.ALL, border=2)
        sizer.Add(boxsizer, pos=(5, 0), span=(1, 5), flag=wx.EXPAND|wx.TOP|wx.LEFT|wx.RIGHT , border=10)

        self.button4 = wx.Button(panel, label="Process")
        sizer.Add(self.button4, pos=(7, 2), span=(1, 1), flag=wx.BOTTOM|wx.RIGHT, border=5)
        self.button4.Disable()
        self.button4.Bind(wx.EVT_BUTTON, self.onProcess)

        self.button5 = wx.Button(panel, label="Abort")
        sizer.Add(self.button5, pos=(7, 3), span=(1, 1), flag=wx.BOTTOM|wx.RIGHT, border=5)
        self.button5.Disable()
        self.button5.Bind(wx.EVT_BUTTON, self.onAbort)
        
        self.button6 = wx.Button(panel, label="Close")
        sizer.Add(self.button6, pos=(7, 4), span=(1, 1), flag=wx.BOTTOM|wx.RIGHT, border=5)
        self.button6.Bind(wx.EVT_BUTTON, self.onClose)
        
        panel.SetSizer(sizer)
        
    def onConfBrowse(self, e):
        dialog = wx.FileDialog(self, message="Choose the Hawk configuration file", wildcard="Configuration Files (*.ini)|*.ini", style=wx.FD_OPEN|wx.FD_FILE_MUST_EXIST)
        if dialog.ShowModal() == wx.ID_OK:
            self.tc1.SetValue(dialog.GetPath())
        dialog.Destroy()
        self.validateFields()
        
    def onSrcBrowse(self, e):
        dialog = wx.FileDialog(self, message="Choose the Source Excel file", wildcard="Excel Files (*.xlsx)|*.xlsx", style=wx.FD_OPEN|wx.FD_FILE_MUST_EXIST)
        if dialog.ShowModal() == wx.ID_OK:
            self.tc2.SetValue(dialog.GetPath())
        dialog.Destroy()
        self.validateFields()
        
    def onDestBrowse(self, e):
        dialog = wx.FileDialog(self, message="Save Results As", wildcard="Excel Files (*.xlsx)|*.xlsx", style=wx.FD_SAVE)
        if dialog.ShowModal() == wx.ID_OK:
            self.tc3.SetValue(dialog.GetPath())
        dialog.Destroy()
        self.validateFields()
        
    def validateFields(self):
        if self.tc1.Value == "[Select File]" or self.tc2.Value == "[Select File]" or self.tc3.Value == "[Specify File]":
            self.button4.Disable()
        else:
            self.button4.Enable()
            
    def onProcess(self, e):
        self.button1.Disable()
        self.button2.Disable()
        self.button3.Disable()
        self.button4.Disable()
        self.button5.Enable()
        self.button6.Disable()
        self.tc4.SetValue('Processing...')
        
        self.thread1 = myThread(self)
        self.thread1.start()
        
    def onAbort(self, e):
        dlg = wx.MessageDialog(self, 'Are you sure about aborting?', 'User Abort', wx.YES_NO | wx.ICON_INFORMATION)
        if dlg.ShowModal() == wx.ID_YES:
            self.thread1.quit()        
            self.tc4.AppendText("\n\n--------------------User Abort--------------------")
            self.button1.Enable()
            self.button2.Enable()
            self.button3.Enable()
            self.button4.Enable()
            self.button5.Disable()
            self.button6.Enable()
        
    def onClose(self, e):
        self.Close(True)
        sys.exit(0)

if __name__ == '__main__':
  
    app = wx.App()
    Hawk(None, title="Hawk 3")
    app.MainLoop()
